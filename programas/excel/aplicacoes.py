# OBJETIVOS 
# - gerar dados aleatorios e:
# - Fazer a media de uma coluna 
# - Achar todas a modas e salvar em outra coluna
# - fazer histograma das modas



from openpyxl import * # importar tds os modulos para planilhas excel
from openpyxl.chart import BarChart, Reference
from myfuncs import *

planilha = Workbook() # cria uma nova planilha
folha = planilha.active # pegar folha ativa

NUMEROS = 1 # na primeira coluna vai ficar os numeros aleatorios
ORDENADOS = 2
MEDIA = 3
MODA = 4
MODA_REPETICOES = 5


numeros_aleatorios = gerarNumerosAleatorios(min = 0,max = 6100, n = 61) # gera 600 numeros aleatorios de 0 ate um 120
numeros_ordenados = numeros_aleatorios[0:]
numeros_ordenados.sort(reverse = True)
mediana = 0

if len(numeros_ordenados)==0:
	mediana = (numeros_ordenados[len(numeros_ordenados)/2] + numeros_ordenados[len(numeros_ordenados)/2-1])/2 
else:
	mediana = numeros_ordenados[int(len(numeros_ordenados)/2)]


length = len(numeros_aleatorios)+1
#colocando lista na planilha
for coluna in folha.iter_cols(min_col = NUMEROS,max_col = NUMEROS,min_row = 2,max_row = length): 
	for celula in coluna:
		celula.value = numeros_aleatorios.pop()# coloca os numeros na primeira coluna
	break

#colocando  segunda lista na planilha
for coluna in folha.iter_cols(min_col = ORDENADOS,max_col = ORDENADOS,min_row = 2,max_row = length): 
	for celula in coluna:
		celula.value = numeros_ordenados.pop()# coloca os numeros na primeira coluna
	break


soma,size,moda = 0,0,{}

for coluna in folha.iter_cols(min_col = NUMEROS,max_col = NUMEROS,min_row = 2,max_row = length): 
	for celula in coluna:
		soma+= float(celula.value)
		size+=1
		if  celula.value in moda.keys():
			moda[celula.value] += 1
		else:
			moda.update({celula.value:1})		
	break

moda = sortDict(moda)
folha.cell(2,MEDIA).value = soma/size # achado a media
folha.cell(4,MEDIA).value = mediana # achado a media
folha.cell(1,MODA).value = 'MODA'
folha.cell(1,MODA_REPETICOES).value = 'REPETIÇÕES'

length_moda = len(moda)
for line in folha.iter_rows(min_col = MODA,max_col = MODA_REPETICOES,min_row = 2,max_row = len(moda)+1): 
	line[0].value,line[1].value = moda.popitem()

	
folha.cell(1,NUMEROS).value = 'Numeros aleatórios'
folha.cell(1,ORDENADOS).value = 'ORDENADOS'
folha.cell(1,MEDIA).value = 'Media'
folha.cell(1,MEDIA).value = 'Media'
folha.cell(3,MEDIA).value = 'Mediana'

#criando grafico de barras
barGraph = BarChart()
data = Reference(worksheet=folha,
                  min_row=1,
                  max_row= length_moda,
                  min_col=MODA,
                  max_col=MODA_REPETICOES)
barGraph.add_data(data,titles_from_data=True)

folha.add_chart(barGraph,'A1')

planilha.save(filename = 'aplicacao.xlsx')





