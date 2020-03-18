#https://realpython.com/openpyxl-excel-spreadsheets-python/#reading-excel-spreadsheets-with-openpyxl
from openpyxl import * # importar tds os modulos
planilha = load_workbook(filename="PYTHON.xlsx") # abre uma planilha ja existente
print(planilha.sheetnames) # mostra os SHEETs existentes

space  = planilha.active # sheet ativo no documento
print(space)
print(space.title)

print(space['A1']) # captura a celula A1
print(space['A1'].value) # mostra o q tem na celula A1

print(space.cell(row  = 1,column =1)) # tambem acessa elementos porem da forma matricial



# iteracao em uma linha,  iter_rows retorna um tuple de tuples correspondente a uma linha
print('\nlinhas:')
for line in space.iter_rows(min_row  = 1,max_row = 1,min_col =1,max_col = 3): 
	print(line)
	
print('\nColunas:')
# iteracao em uma coluna , iter_cols retorna um tuple de tuples cada um correspondendo a uma coluna
for column in space.iter_cols(min_row=1,max_row = 5,min_col =1,max_col = 1):
	print(column)

print('apenas valores')
#deve-se colocar values_only como verdadeiro
for value_column in space.iter_rows(min_row=1,max_row=1,min_col=1,max_col=3,values_only=True):
     print(value_column)


print('percorrente toda coluna ')
for col in space.columns:
    #print(col)
    pass

print('percorrente toda linha ')
for row in space.rows:
    #print(row)
    pass

#editar valor

space.cell(2,1).value = 'anything'

#salvar
planilha.save(filename="PYTHON.xlsx");


#criar nova planilha

novaPlanilha = Workbook()

novaPlanilha.active.cell(1,1).value = "É uma nova planilha :D"
novaPlanilha.save(filename = 'novaPlanilha.xlsx')